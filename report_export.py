# src/report_export.py
# Management reports (Excel + PDF) generated from your existing dataframes.
# English, simple, readable.

from __future__ import annotations

from io import BytesIO
from datetime import datetime
from typing import Optional, Dict, List, Tuple

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# PDF (optional but included)
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table as RLTable, TableStyle
from reportlab.lib.styles import getSampleStyleSheet


# ----------------------------
# Helpers
# ----------------------------

def _safe_col(df: pd.DataFrame, col: str) -> bool:
    return col in df.columns

def _coerce_num(s: pd.Series) -> pd.Series:
    return pd.to_numeric(s, errors="coerce").fillna(0)

def _fmt_int(x) -> int:
    try:
        return int(round(float(x)))
    except Exception:
        return 0

def _alert_norm(x) -> str:
    if x is None:
        return ""
    s = str(x).strip().upper()
    # allow variations
    if "RED" in s:
        return "RED"
    if "ORANGE" in s:
        return "ORANGE"
    if "GREEN" in s:
        return "GREEN"
    return s

def _autofit(ws):
    # simple autofit based on max string length
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row, col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 55)

def _apply_sheet_style(ws):
    font = Font(name="Times New Roman", size=12)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.font = font
            cell.alignment = align
            cell.border = border

    # Freeze header
    ws.freeze_panes = "A2"

def _add_excel_table(ws, table_name: str):
    # add a formatted table with filters
    last_col = get_column_letter(ws.max_column)
    last_row = ws.max_row
    if last_row < 2 or ws.max_column < 1:
        return

    ref = f"A1:{last_col}{last_row}"
    tab = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

def _fill_alert_colors(ws, alert_col_name: str):
    # color cells in the column that equals alert_col_name (header row = 1)
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    headers_norm = [str(h).strip().lower() if h is not None else "" for h in headers]
    try:
        col = headers_norm.index(alert_col_name.strip().lower()) + 1
    except ValueError:
        return

    fill_red = PatternFill("solid", fgColor="FFC7CE")     # light red
    fill_or  = PatternFill("solid", fgColor="FFE699")     # light orange/yellow
    fill_gr  = PatternFill("solid", fgColor="C6EFCE")     # light green

    for r in range(2, ws.max_row + 1):
        v = _alert_norm(ws.cell(r, col).value)
        if v == "RED":
            ws.cell(r, col).fill = fill_red
        elif v == "ORANGE":
            ws.cell(r, col).fill = fill_or
        elif v == "GREEN":
            ws.cell(r, col).fill = fill_gr


# ----------------------------
# Core report builders
# ----------------------------

def build_report_tables(
    master_df: pd.DataFrame,
    plan_df: Optional[pd.DataFrame] = None,
    final_df: Optional[pd.DataFrame] = None,
) -> Dict[str, pd.DataFrame]:
    """
    Returns dict of report tables:
      - executive_summary_kpis (small 2-col table)
      - top10_critical
      - production_priority
      - machine_load
      - final_decision
    """

    master = master_df.copy()

    # Normalize alert
    if _safe_col(master, "Product Alert"):
        master["Product Alert"] = master["Product Alert"].apply(_alert_norm)
    else:
        master["Product Alert"] = ""

    # KPIs
    total_items = int(len(master))

    red = int((master["Product Alert"] == "RED").sum())
    orange = int((master["Product Alert"] == "ORANGE").sum())
    green = int((master["Product Alert"] == "GREEN").sum())

    # Proposed total from plan if available, else 0
    total_proposed = 0
    machines_impacted = 0
    if plan_df is not None and len(plan_df) and _safe_col(plan_df, "Proposed Production Qty"):
        p = plan_df.copy()
        if _safe_col(p, "Machine"):
            machines_impacted = int(p["Machine"].dropna().nunique())
        total_proposed = _coerce_num(p["Proposed Production Qty"]).sum()

    kpis = pd.DataFrame(
        [
            ["Total Items", total_items],
            ["RED Items (Production)", red],
            ["ORANGE Items (Production)", orange],
            ["GREEN Items (Production)", green],
            ["Total Proposed Qty (pcs)", _fmt_int(total_proposed)],
            ["Machines Impacted", machines_impacted],
        ],
        columns=["KPI", "Value"],
    )

    # Build Top 10 Critical Items
    cols_top = []
    for c in ["Item No.", "Item Name", "Machine", "Product Alert", "Total Coverage Days", "Gap Days", "Proposed Production Qty"]:
        if c in master.columns:
            cols_top.append(c)

    # If Machine/Proposed/Gaps are in plan, prefer plan for those
    top_source = master
    if plan_df is not None and len(plan_df):
        # Merge minimal cols from plan onto master by Item No. if possible
        if _safe_col(plan_df, "Item No.") and _safe_col(master, "Item No."):
            pmin_cols = [c for c in ["Item No.", "Machine", "Gap Days", "Proposed Production Qty"] if c in plan_df.columns]
            merged = master.merge(plan_df[pmin_cols].drop_duplicates("Item No."), on="Item No.", how="left", suffixes=("", "_plan"))
            top_source = merged

    # Ensure required columns exist
    if "Total Coverage Days" in top_source.columns:
        top_source["Total Coverage Days"] = pd.to_numeric(top_source["Total Coverage Days"], errors="coerce")
    if "Gap Days" in top_source.columns:
        top_source["Gap Days"] = pd.to_numeric(top_source["Gap Days"], errors="coerce")
    if "Proposed Production Qty" in top_source.columns:
        top_source["Proposed Production Qty"] = pd.to_numeric(top_source["Proposed Production Qty"], errors="coerce").fillna(0)

    # Filter RED/ORANGE first, then sort
    critical = top_source[top_source["Product Alert"].isin(["RED", "ORANGE"])].copy()
    sort_cols = []
    if "Product Alert" in critical.columns:
        # RED before ORANGE
        critical["__alert_rank"] = critical["Product Alert"].map({"RED": 0, "ORANGE": 1}).fillna(9)
        sort_cols.append("__alert_rank")
    if "Gap Days" in critical.columns:
        sort_cols.append("Gap Days")
    if "Total Coverage Days" in critical.columns:
        sort_cols.append("Total Coverage Days")

    if sort_cols:
        # Gap Days descending (bigger gap is worse), Coverage ascending
        ascending = [True] + ([False] if "Gap Days" in sort_cols else []) + ([True] if "Total Coverage Days" in sort_cols else [])
        # fix length mismatch
        while len(ascending) < len(sort_cols):
            ascending.append(True)
        if "Gap Days" in sort_cols:
            # ensure correct ascending list mapping
            # (__alert_rank asc), (Gap Days desc), (Coverage asc)
            ascending = [True]
            if "Gap Days" in sort_cols:
                ascending.append(False)
            if "Total Coverage Days" in sort_cols:
                ascending.append(True)
        critical = critical.sort_values(sort_cols, ascending=ascending)

    # Columns to show
    top_cols_show = [c for c in ["Item No.", "Item Name", "Machine", "Product Alert", "Total Coverage Days", "Gap Days", "Proposed Production Qty"] if c in critical.columns]
    top10 = critical[top_cols_show].head(10).copy()
    if "__alert_rank" in top10.columns:
        top10 = top10.drop(columns=["__alert_rank"])

    # Production Priority table (full, sorted)
    pr_cols = [c for c in ["Item No.", "Item Name", "Machine", "Product Alert", "Monthly Demand", "Total Coverage Days", "Gap Days", "Proposed Production Qty"] if c in top_source.columns]
    production_priority = top_source.copy()
    if "Product Alert" in production_priority.columns:
        production_priority["__alert_rank"] = production_priority["Product Alert"].map({"RED": 0, "ORANGE": 1, "GREEN": 2}).fillna(9)
        sort_cols2 = ["__alert_rank"]
        asc2 = [True]
        if "Gap Days" in production_priority.columns:
            sort_cols2.append("Gap Days"); asc2.append(False)
        if "Total Coverage Days" in production_priority.columns:
            sort_cols2.append("Total Coverage Days"); asc2.append(True)
        production_priority = production_priority.sort_values(sort_cols2, ascending=asc2)
    production_priority = production_priority[pr_cols].copy()
    if "__alert_rank" in production_priority.columns:
        production_priority = production_priority.drop(columns=["__alert_rank"])

    # Machine Load summary
    machine_load = pd.DataFrame()
    if plan_df is not None and len(plan_df) and _safe_col(plan_df, "Machine"):
        p = plan_df.copy()
        if _safe_col(p, "Proposed Production Qty"):
            p["Proposed Production Qty"] = _coerce_num(p["Proposed Production Qty"])
        if _safe_col(p, "Product Alert"):
            p["Product Alert"] = p["Product Alert"].apply(_alert_norm)
        agg = {
            "Item No.": "count" if _safe_col(p, "Item No.") else "size",
        }
        if _safe_col(p, "Proposed Production Qty"):
            agg["Proposed Production Qty"] = "sum"
        if _safe_col(p, "Product Alert"):
            # counts for RED/ORANGE
            p["is_red"] = (p["Product Alert"] == "RED").astype(int)
            p["is_orange"] = (p["Product Alert"] == "ORANGE").astype(int)
            agg["is_red"] = "sum"
            agg["is_orange"] = "sum"

        machine_load = p.groupby("Machine", dropna=True).agg(agg).reset_index()
        # Rename columns
        ren = {}
        if "Item No." in machine_load.columns:
            ren["Item No."] = "Items Count"
        if "Proposed Production Qty" in machine_load.columns:
            ren["Proposed Production Qty"] = "Total Proposed Qty (pcs)"
        if "is_red" in machine_load.columns:
            ren["is_red"] = "RED Count"
        if "is_orange" in machine_load.columns:
            ren["is_orange"] = "ORANGE Count"
        machine_load = machine_load.rename(columns=ren)
        # Sort by proposed qty desc if available
        if "Total Proposed Qty (pcs)" in machine_load.columns:
            machine_load = machine_load.sort_values("Total Proposed Qty (pcs)", ascending=False)

    # Final decision table
    final_decision = pd.DataFrame()
    if final_df is not None and len(final_df):
        f = final_df.copy()
        show = [c for c in ["Item No.", "Item Name", "Machine", "Proposed Production Qty", "Final Qty"] if c in f.columns]
        if "Final Qty" in f.columns:
            f["Final Qty"] = _coerce_num(f["Final Qty"])
        if "Proposed Production Qty" in f.columns:
            f["Proposed Production Qty"] = _coerce_num(f["Proposed Production Qty"])
        final_decision = f[show].copy()

    return {
        "Executive Summary": kpis,
        "Top 10 Critical": top10,
        "Production Priority": production_priority,
        "Machine Load": machine_load,
        "Final Decision": final_decision,
    }


def export_management_report_excel_bytes(
    master_df: pd.DataFrame,
    plan_df: Optional[pd.DataFrame] = None,
    final_df: Optional[pd.DataFrame] = None,
) -> bytes:
    """
    Creates an Excel report with 4-5 sheets:
      - Executive Summary
      - Top 10 Critical
      - Production Priority
      - Machine Load
      - Final Decision
    """
    tables = build_report_tables(master_df, plan_df, final_df)

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for idx, (sheet_name, df) in enumerate(tables.items(), start=1):
        ws = wb.create_sheet(title=sheet_name[:31])

        if df is None or df.empty:
            ws.append(["No data"])
            _apply_sheet_style(ws)
            _autofit(ws)
            continue

        # write header
        ws.append(list(df.columns))
        # write rows
        for row in df.itertuples(index=False):
            ws.append(list(row))

        _apply_sheet_style(ws)

        # Add table (avoid on Executive Summary small kpis if you want — but safe)
        table_name = f"T{idx}_{sheet_name.replace(' ', '')}".replace("-", "")
        _add_excel_table(ws, table_name[:28])

        # Alert colors if sheet has Product Alert
        _fill_alert_colors(ws, "Product Alert")

        _autofit(ws)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def export_management_report_pdf_bytes(
    master_df: pd.DataFrame,
    plan_df: Optional[pd.DataFrame] = None,
    final_df: Optional[pd.DataFrame] = None,
    title: str = "Stock Planner - Management Summary",
) -> bytes:
    """
    Simple 2-page style PDF:
      - Page 1: KPIs + Top 10 Critical
      - Page 2: Production Priority (Top 30) + Machine Load
    """
    tables = build_report_tables(master_df, plan_df, final_df)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, title=title)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(title, styles["Title"]))
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"]))
    story.append(Spacer(1, 10))

    # KPIs
    kpis = tables.get("Executive Summary", pd.DataFrame())
    if not kpis.empty:
        story.append(Paragraph("Executive Summary (KPIs)", styles["Heading2"]))
        data = [list(kpis.columns)] + kpis.values.tolist()
        t = RLTable(data, hAlign="LEFT")
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTNAME", (0, 0), (-1, -1), "Times-Roman"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(t)
        story.append(Spacer(1, 12))

    # Top 10 critical
    top10 = tables.get("Top 10 Critical", pd.DataFrame())
    if not top10.empty:
        story.append(Paragraph("Top 10 Critical Items (Production)", styles["Heading2"]))
        data = [list(top10.columns)] + top10.values.tolist()
        t = RLTable(data, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTNAME", (0, 0), (-1, -1), "Times-Roman"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(t)
        story.append(Spacer(1, 14))

    # Page break-ish
    story.append(Spacer(1, 18))
    story.append(Paragraph("Production Priority (Top 30)", styles["Heading2"]))

    priority = tables.get("Production Priority", pd.DataFrame())
    if not priority.empty:
        top30 = priority.head(30).copy()
        data = [list(top30.columns)] + top30.values.tolist()
        t = RLTable(data, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTNAME", (0, 0), (-1, -1), "Times-Roman"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(t)
        story.append(Spacer(1, 12))
    else:
        story.append(Paragraph("No priority data.", styles["Normal"]))
        story.append(Spacer(1, 12))

    machine_load = tables.get("Machine Load", pd.DataFrame())
    story.append(Paragraph("Machine Load Summary", styles["Heading2"]))
    if not machine_load.empty:
        data = [list(machine_load.columns)] + machine_load.values.tolist()
        t = RLTable(data, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTNAME", (0, 0), (-1, -1), "Times-Roman"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(t)
    else:
        story.append(Paragraph("No machine load data.", styles["Normal"]))

    doc.build(story)
    return buf.getvalue()
