# src/production_plan_export.py
from io import BytesIO
import pandas as pd
import openpyxl


def _digits_key(x) -> str:
    """Digits only + strip leading zeros."""
    if x is None:
        return ""
    s = "".join(ch for ch in str(x) if ch.isdigit())
    return s.lstrip("0")


def fill_qty_in_client_orders(
    template_bytes: bytes,
    plan_df: pd.DataFrame,
    plan_itemno_col: str = "Item No.",
    plan_qty_col: str = "Final Qty",      # أو "Proposed Production Qty"
    sheet_name: str = "Clinet Orders",    # نفس اسم الشيت عندك
    header_row: int = 2,                  # صف الهيدر
    start_row: int = 3,                   # أول صف بيانات
    sheet_itemno_header: str = "Item No.",
    sheet_qty_header: str = "Qty",
) -> tuple[bytes, int, int, list]:
    """
    Fills Qty column in 'Clinet Orders' by matching Item No.
    Writes Qty EVEN IF ZERO.
    Returns: (filled_file_bytes, filled_count, not_matched_count, unmatched_itemnos)
    """

    # --- Validation ---
    if plan_itemno_col not in plan_df.columns:
        raise ValueError(f"plan_df missing column: {plan_itemno_col}")
    if plan_qty_col not in plan_df.columns:
        raise ValueError(f"plan_df missing column: {plan_qty_col}")

    # --- Prepare plan map ---
    df = plan_df[[plan_itemno_col, plan_qty_col]].copy()
    df[plan_itemno_col] = df[plan_itemno_col].apply(_digits_key)
    df[plan_qty_col] = pd.to_numeric(df[plan_qty_col], errors="coerce").fillna(0)

    # Map: ItemNoKey -> Qty (ZERO INCLUDED)
    qty_map = dict(zip(df[plan_itemno_col], df[plan_qty_col]))

    # --- Load template ---
    wb = openpyxl.load_workbook(BytesIO(template_bytes))
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}")

    ws = wb[sheet_name]

    # --- Find columns by headers ---
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    headers_norm = [str(h).strip().lower() if h is not None else "" for h in headers]

    try:
        col_item = headers_norm.index(str(sheet_itemno_header).strip().lower()) + 1
    except ValueError:
        raise ValueError(f"Header '{sheet_itemno_header}' not found in row {header_row}")

    try:
        col_qty = headers_norm.index(str(sheet_qty_header).strip().lower()) + 1
    except ValueError:
        raise ValueError(f"Header '{sheet_qty_header}' not found in row {header_row}")

    # --- Fill Qty ---
    filled = 0
    not_matched = 0
    unmatched = []

    for r in range(start_row, ws.max_row + 1):
        item_val = ws.cell(r, col_item).value
        key = _digits_key(item_val)
        if not key:
            continue

        if key in qty_map:
            q = float(qty_map[key])
            ws.cell(r, col_qty).value = q    # WRITE EVEN IF ZERO
            filled += 1
        else:
            not_matched += 1
            unmatched.append(key)

    # --- Save output ---
    out = BytesIO()
    wb.save(out)

    return out.getvalue(), filled, not_matched, unmatched
