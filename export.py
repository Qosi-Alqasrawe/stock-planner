# src/export.py
from io import BytesIO
import pandas as pd

from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


FILLS = {
    "RED": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
    "ORANGE": PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
    "YELLOW": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
    "GREEN": PatternFill(start_color="00B050", end_color="00B050", fill_type="solid"),
}

BASE_FONT = Font(name="Times New Roman", size=12)
HEADER_FONT = Font(name="Times New Roman", size=12, bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autofit_columns(ws, min_width=10, max_width=60):
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col_idx).value
            if val is None:
                continue
            s = str(val)
            if len(s) > max_len:
                max_len = len(s)
        width = max(min_width, min(max_width, max_len + 2))
        ws.column_dimensions[col_letter].width = width


def _apply_global_format(ws):
    # Freeze header
    ws.freeze_panes = "A2"

    # Set row height (اختياري)
    ws.row_dimensions[1].height = 22

    # Apply font + alignment
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = HEADER_FONT if r == 1 else BASE_FONT
            cell.alignment = CENTER


def _add_table_with_filter(ws, table_name="MasterTable"):
    # نطاق الجدول
    last_col = get_column_letter(ws.max_column)
    last_row = ws.max_row
    ref = f"A1:{last_col}{last_row}"

    tab = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)


def to_excel_bytes(
    master_df: pd.DataFrame,
    machine_df: pd.DataFrame | None,
    itemno_col: str = "Item No.",
    master_sheet: str = "Master",
) -> bytes:
    bio = BytesIO()
    master = master_df.copy()

    # Item No كنص
    if itemno_col in master.columns:
        master[itemno_col] = master[itemno_col].astype(str)

    # خلي مدخلات الخطة فاضية (Excel input)
    for c in ["Plan Qty Input", "Plan Months Input"]:
        if c in master.columns:
            master[c] = ""

    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        master.to_excel(writer, sheet_name=master_sheet, index=False)

        if machine_df is not None and len(machine_df) > 0:
            machine_df.to_excel(writer, sheet_name="Machine_Plan", index=False)

        wb = writer.book

        # ===== Format Master =====
        ws = wb[master_sheet]
        header = [cell.value for cell in ws[1]]

        # Item No as text in Excel
        if itemno_col in header:
            col_idx = header.index(itemno_col) + 1
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col_idx).number_format = "@"

        # Alerts coloring
        def paint(col_name: str):
            if col_name not in header:
                return
            col_idx = header.index(col_name) + 1
            for r in range(2, ws.max_row + 1):
                val = ws.cell(row=r, column=col_idx).value
                fill = FILLS.get(val)
                if fill:
                    ws.cell(row=r, column=col_idx).fill = fill

        paint("Product Alert")
        paint("CUST Alert")

        # Excel formulas (Months Covered / Qty Needed)
        needed_cols = [
            "Plan Qty Input",
            "Months Covered by Plan Qty",
            "Plan Months Input",
            "Qty Needed for Plan Months",
            "Monthly Demand",
        ]
        if all(c in header for c in needed_cols):
            col_plan_qty = header.index("Plan Qty Input") + 1
            col_months_cov = header.index("Months Covered by Plan Qty") + 1
            col_plan_months = header.index("Plan Months Input") + 1
            col_qty_needed = header.index("Qty Needed for Plan Months") + 1
            col_md = header.index("Monthly Demand") + 1

            for r in range(2, ws.max_row + 1):
                pq = f"{get_column_letter(col_plan_qty)}{r}"
                md = f"{get_column_letter(col_md)}{r}"
                pm = f"{get_column_letter(col_plan_months)}{r}"

                ws.cell(row=r, column=col_months_cov).value = f'=IFERROR({pq}/{md},0)'
                ws.cell(row=r, column=col_qty_needed).value = f'=IFERROR({pm}*{md},0)'

                ws.cell(row=r, column=col_months_cov).number_format = "0.0"
                ws.cell(row=r, column=col_qty_needed).number_format = "0"

        # Apply global formatting + table + autofit
        _apply_global_format(ws)
        _add_table_with_filter(ws, table_name="MasterTable")
        _autofit_columns(ws)

        # ===== Optional: format Machine_Plan =====
        if machine_df is not None and len(machine_df) > 0 and "Machine_Plan" in wb.sheetnames:
            ws2 = wb["Machine_Plan"]
            _apply_global_format(ws2)
            _add_table_with_filter(ws2, table_name="MachinePlanTable")
            _autofit_columns(ws2)

    return bio.getvalue()

# ✅ NEW: export multiple sheets (Sheet per machine)
def to_excel_bytes_multi_sheets(
    sheets: dict[str, pd.DataFrame],
    itemno_col: str = "Item No.",
) -> bytes:
    """
    Export one Excel file with multiple sheets.
    sheets: {"SheetName": dataframe, ...}
    Applies same formatting (Times New Roman 12, center, wrap, table+filters, autofit, freeze header)
    and colors Product Alert / CUST Alert if present.
    """
    bio = BytesIO()

    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        # write all sheets
        for sheet_name, df in sheets.items():
            if df is None or len(df) == 0:
                continue

            out = df.copy()
            if itemno_col in out.columns:
                out[itemno_col] = out[itemno_col].astype(str)

            out.to_excel(writer, sheet_name=sheet_name, index=False)

        wb = writer.book

        # format each sheet
        for idx, sheet_name in enumerate(wb.sheetnames, start=1):
            ws = wb[sheet_name]
            header = [cell.value for cell in ws[1]]

            # Item No as text in Excel
            if itemno_col in header:
                col_idx = header.index(itemno_col) + 1
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=col_idx).number_format = "@"

            # Alerts coloring (if exist)
            def paint(col_name: str):
                if col_name not in header:
                    return
                col_idx = header.index(col_name) + 1
                for r in range(2, ws.max_row + 1):
                    val = ws.cell(row=r, column=col_idx).value
                    fill = FILLS.get(val)
                    if fill:
                        ws.cell(row=r, column=col_idx).fill = fill

            paint("Product Alert")
            paint("CUST Alert")

            # Global format + table + autofit
            _apply_global_format(ws)

            # unique table name per sheet (Excel requires unique names)
            table_name = f"Tbl{idx}"
            _add_table_with_filter(ws, table_name=table_name)

            _autofit_columns(ws)

    return bio.getvalue()
